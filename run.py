import re, sys, os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import win32com.client as win32
import xlwings as xw
from webtrax.webtrax import Webtrax


def update_excel(data_frame, file_name, sh_name):
    """hello"""
    xlbook = load_workbook(file_name)
    xlsheet = xlbook[sh_name]
    for row in xlsheet.rows:
        for cell in row:
            cell.value = None
    rows = dataframe_to_rows(data_frame, index=False, header=True)
    for row_num, row in enumerate(rows, 1):
        for col_num, value in enumerate(row, 1):
            xlsheet.cell(row_num, col_num, value)
    xlbook.save(file_name)

def run_excel_macro(sfile_path, ssheet_name, smacro_path, df_data):
    """hello"""
    win32.pythoncom.CoInitialize()
    xl_app  = win32.Dispatch('Excel.Application')
    xl_app.Visible = True
    xlwb = xl_app.Workbooks.Open(sfile_path)
    xlsh = xlwb.Worksheets(ssheet_name)
    rng = xlsh.Range("A1").CurrentRegion
    rng.ClearContents()
    StartRow = 2
    StartCol = 1
    xlsh.Range(xlsh.Cells(StartRow-1,StartCol), xlsh.Cells(StartRow-1,StartCol+len(df_data.columns)-1)).Value = df_data.columns.values
    xlsh.Range(xlsh.Cells(StartRow,StartCol), xlsh.Cells(StartRow+len(df_data.index)-1,StartCol+len(df_data.columns)-1)).Value = df_data.values
    xl_app.Application.Run(smacro_path)

def update_and_run_excel(sfile_path, smacroname, data_frame):
    """test"""
    wb = xw.Book(sfile_path)
    ws = wb.sheets['Inventory']
    ws.range('A2:P2000').clear_contents()
    ws.range('A2').options(index=False, header=False).value = data_frame
    run = wb.macro(smacroname)
    run()
    wb.save()
    wb.close()

def share_report(filename, macroname):
    """share report"""
    wb = xw.Book(filename)
    run = wb.macro(macroname)
    run()
    wb.close()

with Webtrax(teardown = False) as bot:
    file_path = bot.get_file_path(r"Webtrax Escalation Inventory.xlsm")
    sheet_name = r"Inventory"
    macro_path=r'mod_Main.AutoReport'
    bot.land_page()
    data_df = pd.DataFrame()
    url_list = bot.get_links()
    count = 0
    for url in url_list:
        bot.navigate_to_page(url)
        df = bot.get_datatable()
        queue = re.findall(r"(?<=&Que=).+", url) * df['Age'].count()
        location = re.findall(r"(?<=&Type=)(.+)&", url) * df['Age'].count()
        df.insert(0,'Queue', queue)
        df.insert(1,'Location', location)
        data_df = pd.concat([data_df, df])
    # run_excel_macro(file_path, sheet_name, macro_path, data_df)
    base_path = os.path.join(re.split(os.environ.get('username'), sys.executable)[0], os.environ.get('username'))
    file_path = os.path.join(base_path, 'Documents', 'Templates', 'Webtrax Escalation Inventory.xlsm')
    print(file_path)
    update_and_run_excel(file_path, macro_path, data_df)
    bot.quit()
    # update_excel(data_df, r"Webtrax Escalation Inventory.xlsx", 'Inventory')
    # share_report('Webtrax Escalation Inventory.xlsx', "mod_Main.AutoReport")